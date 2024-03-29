import itertools
import os
from pathlib import Path
import random

import numpy as np
from openpyxl import load_workbook
import tensorflow as tf
import keras

def load(image_path, *args, **kwargs):
    image = keras.utils.load_img(image_path, *args, **kwargs)
    image = np.array(image)
    return image

def get_exposure_dataset_item(filename):
    # Load image
    image = load(filename, target_size=(256, 256), interpolation='bilinear')

    # Parse label into float - labels in filename are either N<number>, 0, or P<number>
    label_string = filename.split(b'_')[-1].rsplit(b'.', 1)[0]
    if label_string[0] == ord('P'):
        label = float(label_string[1:])
    elif label_string[0] == ord('N'):
        label = -float(label_string[1:])
    else:
        assert label_string == b'0', f'Unexpected label: {label_string}'
        label = 0.0

    return image, label

def generate_exposure_dataset(filenames):
    for filename in filenames:
        yield get_exposure_dataset_item(filename)

def load_exposure_dataset(folders):
    label_files = []
    for folder in folders:
        label_files.extend(os.path.join(folder, f) for f in os.listdir(folder))
    dataset = tf.data.Dataset.from_generator(
        generate_exposure_dataset,
        args=(label_files,),
        output_signature=(
            tf.TensorSpec(shape=(256, 256, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset

def get_blur_dataset_item(filename, label):
    # Load image
    image = load(filename)

    return image, label

def generate_realblur_dataset(list_file):
    base_dir = os.path.dirname(list_file)
    with open(list_file, 'r') as f:
        for line in f.readlines():
            gt_filename, blurred_filename, *_ = line.split(' ')
            yield get_blur_dataset_item(os.path.join(base_dir, blurred_filename.encode('utf-8')), 1)
            yield get_blur_dataset_item(os.path.join(base_dir, gt_filename.encode('utf-8')), 0)

def load_realblur_dataset(list_file):
        dataset = tf.data.Dataset.from_generator(
            generate_realblur_dataset,
            args=(str(list_file),),
            output_signature=(
                tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
                tf.TensorSpec(shape=(), dtype=tf.float32),
            ),
        )

        return dataset

def generate_blur_type_dataset(base_folder):
    base_folder = Path(base_folder.decode('utf-8'))

    label_key = {
        'sharp': 0,
        'motion_blurred': 1,
        'defocused_blurred': 2,
    }

    samples = list(Path(base_folder, 'sharp').iterdir())
    samples.extend(Path(base_folder, 'motion_blurred').iterdir())
    samples.extend(Path(base_folder, 'defocused_blurred').iterdir())
    random.seed(0)
    random.shuffle(samples)
    for image_file in samples:
        image = load(image_file)
        label = label_key[image_file.parent.name]

        yield image, label

def load_blur_type_dataset(folder):
    dataset = tf.data.Dataset.from_generator(
        generate_blur_type_dataset,
        args=(str(folder),),
        output_signature=(
            tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset

def generate_certh_training_dataset(base_folder):
    folders = {
        os.path.join(base_folder, b'Undistorted'): 0,
        os.path.join(base_folder, b'Artificially-Blurred'): 1,
        os.path.join(base_folder, b'Naturally-Blurred'): 1,
    }
    files = [(os.path.join(folder, filename), label) for folder, label in folders.items() for filename in os.listdir(folder)]
    random.shuffle(files)
    for filename, label in files:
        yield get_blur_dataset_item(filename, label)

def load_certh_training_dataset(folder):
    dataset = tf.data.Dataset.from_generator(
        generate_certh_training_dataset,
        args=(str(folder),),
        output_signature=(
            tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset

def generate_certh_testing_dataset(folder):
    folder = folder.decode('utf-8')
    label_folder, label_file = os.path.split(folder)
    label_file = f'{label_file}.xlsx'
    label_file_path = os.path.join(label_folder, label_file)
    labels = load_workbook(label_file_path, read_only=True).get_sheet_by_name('Φύλλο1')
    for name, label in labels.iter_rows(min_row=2):
        filename = os.path.join(folder, f'{name.value}.jpg')
        label = 1 if label.value == 1 else 0
        yield get_blur_dataset_item(filename, label)

def load_certh_testing_dataset(folder):
    dataset = tf.data.Dataset.from_generator(
        generate_certh_testing_dataset,
        args=(str(folder),),
        output_signature=(
            tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset

def generate_sidd_dataset(folder):
    folder = folder.decode('utf-8')
    for subfolder in os.listdir(folder):
        subfolder_path = os.path.join(folder, subfolder)
        for filename in os.listdir(subfolder_path):
            _, label_string, *_ = filename.split('_')
            image = load(os.path.join(subfolder_path, filename))
            label = 1 if label_string == 'NOISY' else 0
            yield image, label

def load_sidd_dataset(folder):
    dataset = tf.data.Dataset.from_generator(
        generate_sidd_dataset,
        args=(str(folder),),
        output_signature=(
            tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset

def generate_ebb_dataset(root_folder):
    root_folder = root_folder.decode('utf-8')
    # Determine files in dataset, then deterministically shuffle
    files = list(Path(root_folder, 'bokeh').iterdir())
    files.extend(Path(root_folder, 'original').iterdir())
    random.seed(0)
    random.shuffle(files)

    for file in files:
        image = load(file)
        label = 1 if file.parent.name == 'bokeh' else 0
        yield image, label

def load_ebb_dataset(root_folder):
    dataset = tf.data.Dataset.from_generator(
        generate_ebb_dataset,
        args=(str(root_folder),),
        output_signature=(
            tf.TensorSpec(shape=(None, None, 3), dtype=tf.float32),
            tf.TensorSpec(shape=(), dtype=tf.float32),
        ),
    )

    return dataset
